import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np
from autolab_core import YamlConfig
plt.rcParams['font.size']=21#35
cfg = YamlConfig("cfg/plot_planner_data.yaml")
#workbook = load_workbook("data/Experiments for revision.xlsx")
workbook = load_workbook(cfg["workbook_path"])
data_type = cfg["data_type"]
eps = cfg["eps"]
MONOSPACE_FONT_DICT = {"family":"monospace", "weight": "bold"}
task_names = ["RodInBox", "RodInDrawer"]
#methods = [anchor_baseline, random_baseline, simulator_only, analytical_only_rod_and_robot, analytical_only_rod_and_drawer, ours]

method_data = {}
method_names = ["anchor_baseline", "ours", "random_baseline", "sim_only", "a_rod_only", "aonly_drawer"]
for method_name in method_names:
    method_data_per_method_name = {}
    for task_name in task_names:
        sheet_name = f"{method_name}_{task_name}_real"
        method_data_per_method_name["title"] = cfg["per_method_info"][method_name]["title"]
        try:
            relevant_sheet = workbook[sheet_name]
        except:
            sheet_name = f"{method_name}_{task_name}_rea" #bug somewhere, probably due to long sheet names
            relevant_sheet = workbook[sheet_name]

        plan_found_data = [relevant_sheet.cell(row=i, column = 3).value for i in range(6,16)]

        plan_success_data = [relevant_sheet.cell(row=i, column = 4).value for i in range(6,16)]
        plan_success_data = [elt for i, elt in enumerate(plan_success_data) if elt != "n/a" and plan_found_data[i]]
        for data_list in [plan_found_data, plan_success_data]:
            if sum(data_list) == 0:
                for item in data_list:
                    item += eps #makes the plot look a little nicer
        method_data_per_method_name[task_name] = {}
        method_data_per_method_name[task_name]["plan_success"] = plan_success_data
        method_data_per_method_name[task_name]["plan_found"] = plan_found_data
    method_data[method_name] = method_data_per_method_name




if cfg.get("plot_one_task_only", False):
    X  = np.arange(1)
    task_names = [cfg["task_for_bar_graph"]]
else:
    X = np.arange(2)

fig = plt.figure()
ax = fig.add_subplot(111)
for i, method_name in enumerate(method_names):
    mean_scores = [max(eps, np.mean(method_data[method_name][task][data_type])) for task in task_names]
    rects = ax.bar(X + i*(cfg["bar_width"]+cfg["bar_spacing"]), mean_scores, width=cfg["bar_width"], label=method_data[method_name]['title'])
    ax.bar_label(ax.containers[i], fmt="%.2f")

if cfg.get("plot_one_task_only", False):
    ax.set_xticks([0 + i*(cfg["bar_width"]+cfg["bar_spacing"]) for i in range(len(method_names))])
    ax.set_xlabel("Method")
    ax.set_xticklabels([method_data[method_name]['title'] for method_name in method_names])
    #plt.title(task_names[0], fontdict=MONOSPACE_FONT_DICT)

else:
    len_bars_for_task  = len(method_names)*(cfg['bar_width'] + cfg["bar_spacing"])
    ax.set_xticks([x + len_bars_for_task/2 - cfg['bar_width']/2  - cfg["bar_spacing"]/2 for x in X])
    ax.set_xticklabels(task_names, fontdict=MONOSPACE_FONT_DICT)
    ax.set_xlabel("Task")
    #plt.legend()

if data_type == "plan_success":
    ax.set_ylabel("Plan execution success")
else:
    ax.set_ylabel("Plan found")
ax.set_ylim([0,1])
#plt.subplots_adjust(right=0.8)
plt.subplots_adjust(bottom=cfg.get('bottom_space',0.4))
plt.show()


